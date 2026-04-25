from django.shortcuts import render, redirect, get_object_or_404
from .models import Task, Comment, TaskChangeLog
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .forms import TaskCreateForm, CommentCreateForm
from projects.models import Project
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def format_field_value(field_name, value):
    if value is None:
        return ''

    if field_name == 'status':
        status_map = {
            'todo': 'To Do',
            'in_progress': 'In Progress',
            'done': 'Done',
        }
        return status_map.get(value, str(value))

    if field_name in ['executor', 'project']:
        return str(value)

    return str(value)

def log_task_changes(task, old_task, changed_by):
    tracked_fields = ['title', 'description', 'due_date', 'status', 'executor', 'project']

    for field in tracked_fields:
        old_value = getattr(old_task, field)
        new_value = getattr(task, field)

        if old_value != new_value:
            TaskChangeLog.objects.create(
                task=task,
                changed_by=changed_by,
                field_name=field,
                old_value=format_field_value(field, old_value),
                new_value=format_field_value(field, new_value),
            )

def log_watchers_changes(task, old_watchers, changed_by):
    new_watchers = set(task.watchers.all())

    old_ids = {user.id for user in old_watchers}
    new_ids = {user.id for user in new_watchers}

    if old_ids != new_ids:
        old_value = ', '.join(user.username for user in old_watchers)
        new_value = ', '.join(user.username for user in new_watchers)

        TaskChangeLog.objects.create(
            task=task,
            changed_by=changed_by,
            field_name='watchers',
            old_value=old_value,
            new_value=new_value,
        )

def task_list(request):
    if not request.user.is_authenticated:
        return render(request, 'tasks/task_list.html', {
            'is_guest': True,
        })

    tasks = Task.objects.filter(
        Q(executor=request.user) |
        Q(creator=request.user) |
        Q(watchers=request.user)
    ).select_related(
        'creator',
        'executor',
        'project',
    ).prefetch_related(
        'watchers',
    ).distinct()

    user_projects = Project.objects.filter(
        Q(tasks__executor=request.user) |
        Q(tasks__creator=request.user) |
        Q(tasks__watchers=request.user)
    ).distinct().order_by('name')

    selected_project_id = request.GET.get('project')

    if selected_project_id:
        tasks = tasks.filter(project_id=selected_project_id)

    todo_tasks = list(tasks.filter(status='todo'))
    in_progress_tasks = list(tasks.filter(status='in_progress'))
    done_tasks = list(tasks.filter(status='done'))

    for task_list_group in (todo_tasks, in_progress_tasks, done_tasks):
        for task in task_list_group:
            if task.creator_id == request.user.id:
                task.user_role_label = 'Вы постановщик'
                task.user_role_class = 'creator'
            elif task.executor_id == request.user.id:
                task.user_role_label = 'Вы исполнитель'
                task.user_role_class = 'executor'
            elif any(w.id == request.user.id for w in task.watchers.all()):
                task.user_role_label = 'Вы наблюдатель'
                task.user_role_class = 'watcher'
            else:
                task.user_role_label = ''
                task.user_role_class = ''

    return render(request, 'tasks/task_list.html', {
        'is_guest': False,
        'todo_tasks': todo_tasks,
        'in_progress_tasks': in_progress_tasks,
        'done_tasks': done_tasks,
        'projects': user_projects,
        'selected_project_id': selected_project_id,
    })

@login_required
def task_create(request):
    if request.method == 'POST':
        form = TaskCreateForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.creator = request.user
            task.save()
            form.save_m2m()
            tracked_fields = ['title', 'description', 'due_date', 'status', 'executor', 'project']
            for field in tracked_fields:
                value = getattr(task, field)
                if value:
                    TaskChangeLog.objects.create(
                        task=task,
                        changed_by=request.user,
                        field_name=field,
                        old_value='',
                        new_value=format_field_value(field, value),
                    )

            watchers_value = ', '.join(user.username for user in task.watchers.all())
            if watchers_value:
                TaskChangeLog.objects.create(
                    task=task,
                    changed_by=request.user,
                    field_name='watchers',
                    old_value='',
                    new_value=watchers_value,
                )
            return redirect('tasks:list')
    else:
        form = TaskCreateForm(user=request.user)

    return render(request, 'tasks/task_form.html', {'form': form, 'button_text': 'Создать задачу',})

@login_required
def task_detail(request, pk):
    task = get_object_or_404(
        Task.objects.select_related(
            'creator', 'executor', 'project'
        ).prefetch_related(
            'watchers',
            'comments__user'
        ).distinct(),
        Q(pk=pk) & (
            Q(executor=request.user) |
            Q(creator=request.user) |
            Q(watchers=request.user)
        )
    )

    can_edit = request.user == task.creator or request.user == task.executor
    can_comment = can_edit or task.watchers.filter(pk=request.user.pk).exists()

    if request.method == 'POST':
        if not can_comment:
            return redirect('tasks:detail', pk=task.pk)

        comment_form = CommentCreateForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.task = task
            comment.user = request.user
            comment.save()
            return redirect('tasks:detail', pk=task.pk)
    else:
        comment_form = CommentCreateForm() if can_comment else None

    return render(request, 'tasks/task_detail.html', {
        'task': task,
        'comment_form': comment_form,
        'can_edit': can_edit,
        'can_comment': can_comment,
    })

@login_required
def task_update(request, pk):
    task = get_object_or_404(
        Task.objects.filter(
            Q(executor=request.user) | Q(creator=request.user)
        ),
        pk=pk,
    )

    if request.method == 'POST':
        old_task = Task.objects.get(pk=task.pk)
        old_watchers = list(task.watchers.all())
        form = TaskCreateForm(request.POST, instance=task)
        if form.is_valid():
            updated_task = form.save()
            log_task_changes(updated_task, old_task, request.user)
            log_watchers_changes(updated_task, old_watchers, request.user)
            return redirect('tasks:detail', pk=task.pk)
        
    else:
        form = TaskCreateForm(instance=task, user=request.user)

    return render(request, 'tasks/task_form.html', {
        'form': form,
        'task': task,
        'page_title': 'Редактирование задачи',
        'button_text': 'Сохранить изменения',
    })

@login_required
def comment_update(request, task_pk, comment_pk):
    task = get_object_or_404(
        Task.objects.filter(
            Q(executor=request.user) |
            Q(creator=request.user) |
            Q(watchers=request.user)
        ).distinct(),
        pk=task_pk,
    )

    comment = get_object_or_404(
        Comment,
        pk=comment_pk,
        task=task,
        user=request.user,
    )

    if request.method == 'POST':
        form = CommentCreateForm(request.POST, instance=comment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Комментарий обновлён.')
            return redirect('tasks:detail', pk=task.pk)
    else:
        form = CommentCreateForm(instance=comment)

    return render(request, 'tasks/comment_form.html', {
        'task': task,
        'form': form,
        'page_title': 'Редактирование комментария',
        'button_text': 'Сохранить изменения',
    })


@login_required
def comment_delete(request, task_pk, comment_pk):
    task = get_object_or_404(
        Task.objects.filter(
            Q(executor=request.user) |
            Q(creator=request.user) |
            Q(watchers=request.user)
        ).distinct(),
        pk=task_pk,
    )

    comment = get_object_or_404(
        Comment,
        pk=comment_pk,
        task=task,
        user=request.user,
    )

    if request.method == 'POST':
        comment.delete()
        messages.success(request, 'Комментарий удалён.')
        return redirect('tasks:detail', pk=task.pk)

    return render(request, 'tasks/comment_confirm_delete.html', {
        'task': task,
        'comment': comment,
    })

@login_required
def task_report_xlsx(request, pk):
    task = get_object_or_404(
        Task.objects.select_related(
            'creator',
            'executor',
            'project',
        ).prefetch_related(
            'watchers',
            'change_logs',
        ),
        Q(pk=pk) & (
            Q(creator=request.user) |
            Q(executor=request.user) |
            Q(watchers=request.user)
        )
    )

    workbook = Workbook()

    # ---------------------------
    # Лист 1: общая информация
    # ---------------------------
    task_sheet = workbook.active
    task_sheet.title = 'Task'

    task_sheet['A1'] = 'Поле'
    task_sheet['B1'] = 'Значение'

    task_sheet['A1'].font = Font(bold=True)
    task_sheet['B1'].font = Font(bold=True)

    watchers_value = ', '.join(user.username for user in task.watchers.all()) or 'Нет наблюдателей'
    project_value = task.project.name if task.project else 'Не указан'
    executor_value = task.executor.username if task.executor else 'Не назначен'
    due_date_value = task.due_date.strftime('%d.%m.%Y %H:%M') if task.due_date else 'Не указан'

    task_data = [
        ('ID задачи', task.pk),
        ('Название', task.title),
        ('Описание', task.description or ''),
        ('Статус', task.get_status_display()),
        ('Постановщик', task.creator.username),
        ('Исполнитель', executor_value),
        ('Наблюдатели', watchers_value),
        ('Проект', project_value),
        ('Срок', due_date_value),
        ('Создано', task.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Обновлено', task.updated_at.strftime('%d.%m.%Y %H:%M')),
    ]

    for row_num, (label, value) in enumerate(task_data, start=2):
        task_sheet[f'A{row_num}'] = label
        task_sheet[f'B{row_num}'] = value

    task_sheet.column_dimensions['A'].width = 22
    task_sheet.column_dimensions['B'].width = 50

    # ---------------------------
    # Лист 2: история изменений
    # ---------------------------
    changes_sheet = workbook.create_sheet(title='Changes')

    headers = ['Дата изменения', 'Кто изменил', 'Поле', 'Старое значение', 'Новое значение']
    for col_num, header in enumerate(headers, start=1):
        cell = changes_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    change_logs = task.change_logs.select_related('changed_by').all().order_by('-changed_at')

    for row_num, log in enumerate(change_logs, start=2):
        changes_sheet.cell(row=row_num, column=1).value = log.changed_at.strftime('%d.%m.%Y %H:%M')
        changes_sheet.cell(
            row=row_num,
            column=2
        ).value = log.changed_by.username if log.changed_by else 'Неизвестно'
        changes_sheet.cell(row=row_num, column=3).value = log.field_name
        changes_sheet.cell(row=row_num, column=4).value = log.old_value or ''
        changes_sheet.cell(row=row_num, column=5).value = log.new_value or ''

    column_widths = {
        1: 20,
        2: 20,
        3: 20,
        4: 40,
        5: 40,
    }

    for col_num, width in column_widths.items():
        col_letter = get_column_letter(col_num)
        changes_sheet.column_dimensions[col_letter].width = width

    # ---------------------------
    # Отдача файла
    # ---------------------------
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f'task_report_{task.pk}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response