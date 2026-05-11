from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone

from users.models import User, Department
from tasks.models import Task
from projects.models import Project

from io import BytesIO

from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import timedelta
from django.db import models
import json


@login_required
def reports_view(request):
    now = timezone.now()

    # Отчёт по исполнителям
    users_report = []

    users = User.objects.all()

    for user in users:
        tasks = Task.objects.filter(executor=user, is_archived=False)

        users_report.append({
            "user": user,
            "total": tasks.count(),
            "todo": tasks.filter(status=Task.Status.TODO).count(),
            "in_progress": tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            "completed": tasks.filter(status=Task.Status.DONE).count(),
            "overdue": tasks.filter(
                due_date__lt=now,
                status__in=[Task.Status.TODO, Task.Status.IN_PROGRESS],
            ).count(),
        })

    # Отчёт по проектам
    projects_report = []

    projects = Project.objects.all()

    for project in projects:
        tasks = Task.objects.filter(project=project, is_archived=False)

        projects_report.append({
            "project": project,
            "total": tasks.count(),
            "todo": tasks.filter(status=Task.Status.TODO).count(),
            "in_progress": tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            "completed": tasks.filter(status=Task.Status.DONE).count(),
            "overdue": tasks.filter(
                due_date__lt=now,
                status__in=[Task.Status.TODO, Task.Status.IN_PROGRESS],
            ).count(),
        })

    # Отчёт по отделам
    departments_report = []

    departments = Department.objects.all()

    for department in departments:
        tasks = Task.objects.filter(
            executor__department=department,
            is_archived=False,
        )

        departments_report.append({
            "department": department,
            "users_count": department.users.count(),
            "total": tasks.count(),
            "todo": tasks.filter(status=Task.Status.TODO).count(),
            "in_progress": tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            "completed": tasks.filter(status=Task.Status.DONE).count(),
            "overdue": tasks.filter(
                due_date__lt=now,
                status__in=[Task.Status.TODO, Task.Status.IN_PROGRESS],
            ).count(),
        })

    # Отчёт по нагрузке
    workload_report = []

    users = User.objects.select_related('department').all()

    for user in users:
        tasks = Task.objects.filter(
            executor=user,
            is_archived=False,
        )

        active_tasks = tasks.filter(
            status__in=[
                Task.Status.TODO,
                Task.Status.IN_PROGRESS,
            ]
        )

        active_count = active_tasks.count()

        if active_count >= 10:
            workload_level = "Высокая"
        elif active_count >= 5:
            workload_level = "Средняя"
        else:
            workload_level = "Низкая"

        workload_report.append({
            "user": user,
            "department": user.department.name if user.department else "Без отдела",

            "active": active_count,

            "todo": active_tasks.filter(
                status=Task.Status.TODO
            ).count(),

            "in_progress": active_tasks.filter(
                status=Task.Status.IN_PROGRESS
            ).count(),

            "high_priority": active_tasks.filter(
                priority=Task.Priority.HIGH
            ).count(),

            "overdue": active_tasks.filter(
                due_date__lt=now,
            ).count(),

            "level": workload_level,
        })
    
    # Отчёт по приоритетам
    priority_report = []

    priorities = [
        Task.Priority.HIGH,
        Task.Priority.MEDIUM,
        Task.Priority.LOW,
    ]

    for priority in priorities:
        tasks = Task.objects.filter(
            priority=priority,
            is_archived=False,
        )

        active_tasks = tasks.filter(
            status__in=[
                Task.Status.TODO,
                Task.Status.IN_PROGRESS,
            ]
        )

        priority_report.append({
            "priority": priority,
            "priority_display": Task.Priority(priority).label,

            "total": tasks.count(),

            "active": active_tasks.count(),

            "todo": tasks.filter(
                status=Task.Status.TODO
            ).count(),

            "in_progress": tasks.filter(
                status=Task.Status.IN_PROGRESS
            ).count(),

            "completed": tasks.filter(
                status=Task.Status.DONE
            ).count(),

            "overdue": active_tasks.filter(
                due_date__lt=now,
            ).count(),
        })
    
    priority_chart_labels = []
    priority_chart_data = []

    for item in priority_report:
        priority_chart_labels.append(item['priority_display'])
        priority_chart_data.append(item['total'])

    # Отчёт по скорости команды
    velocity_report = []

    last_7_days = now - timedelta(days=7)
    last_30_days = now - timedelta(days=30)

    users = User.objects.select_related('department').all()

    for user in users:
        completed_tasks = Task.objects.filter(
            executor=user,
            status=Task.Status.DONE,
            completed_at__isnull=False,
            is_archived=False,
        )

        completed_7_days = completed_tasks.filter(
            completed_at__gte=last_7_days,
        ).count()

        completed_30_days = completed_tasks.filter(
            completed_at__gte=last_30_days,
        ).count()

        overdue_completed = completed_tasks.filter(
            due_date__isnull=False,
            completed_at__gt=models.F('due_date'),
        ).count()

        velocity_report.append({
            "user": user,
            "department": user.department.name if user.department else "Без отдела",
            "completed_7_days": completed_7_days,
            "completed_30_days": completed_30_days,
            "completed_total": completed_tasks.count(),
            "overdue_completed": overdue_completed,
        })

    velocity_chart_labels = []
    velocity_chart_data = []

    for i in range(29, -1, -1):
        day = now.date() - timedelta(days=i)

        completed_count = Task.objects.filter(
            status=Task.Status.DONE,
            completed_at__date=day,
            completed_at__isnull=False,
            is_archived=False,
        ).count()

        velocity_chart_labels.append(day.strftime('%d.%m'))
        velocity_chart_data.append(completed_count)


    # SLA отчет по исполнителям
    sla_report = []

    users = User.objects.select_related('department').all()

    for user in users:
        completed_tasks = Task.objects.filter(
            executor=user,
            status=Task.Status.DONE,
            completed_at__isnull=False,
            due_date__isnull=False,
            is_archived=False,
        )

        total_completed = completed_tasks.count()

        completed_on_time = completed_tasks.filter(
            completed_at__lte=models.F('due_date')
        ).count()

        completed_overdue = completed_tasks.filter(
            completed_at__gt=models.F('due_date')
        ).count()

        if total_completed > 0:
            sla_percent = round((completed_on_time / total_completed) * 100, 1)
        else:
            sla_percent = 0

        if sla_percent >= 90:
            sla_level = 'Высокий'
        elif sla_percent >= 70:
            sla_level = 'Средний'
        else:
            sla_level = 'Низкий'

        sla_report.append({
            'user': user,
            'department': user.department.name if user.department else 'Без отдела',
            'total_completed': total_completed,
            'completed_on_time': completed_on_time,
            'completed_overdue': completed_overdue,
            'sla_percent': sla_percent,
            'sla_level': sla_level,
        })

    sla_completed_on_time = Task.objects.filter(
        status=Task.Status.DONE,
        completed_at__isnull=False,
        due_date__isnull=False,
        completed_at__lte=models.F('due_date'),
        is_archived=False,
    ).count()

    sla_completed_overdue = Task.objects.filter(
        status=Task.Status.DONE,
        completed_at__isnull=False,
        due_date__isnull=False,
        completed_at__gt=models.F('due_date'),
        is_archived=False,
    ).count()

    sla_chart_labels = [
        'Вовремя',
        'С просрочкой',
    ]

    sla_chart_data = [
        sla_completed_on_time,
        sla_completed_overdue,
    ]

    return render(request, "reports/reports.html", {
        "users_report": users_report,
        "projects_report": projects_report,
        "departments_report": departments_report,
        "workload_report": workload_report,
        "priority_report": priority_report,
        "priority_chart_labels": priority_chart_labels,
        "priority_chart_data": priority_chart_data,
        "velocity_report": velocity_report,
        "velocity_chart_labels": json.dumps(velocity_chart_labels),
        "velocity_chart_data": json.dumps(velocity_chart_data),
        "sla_report": sla_report,
        "sla_chart_labels": json.dumps(sla_chart_labels),
        "sla_chart_data": json.dumps(sla_chart_data),
    })

def prepare_xlsx_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def style_sheet(sheet):
    header_fill = PatternFill(
        start_color='E5E7EB',
        end_color='E5E7EB',
        fill_type='solid'
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ''
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 60)


def get_task_url(request, task):
    return request.build_absolute_uri(
        reverse('tasks:detail', kwargs={'pk': task.pk})
    )

@login_required
def executors_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Отчёт по исполнителям'

    sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    sheet.append([])

    sheet.append([
        'Исполнитель',
        'Email',
        'Отдел',
        'Всего задач',
        'К выполнению',
        'В работе',
        'Выполнено',
        'Просрочено',
    ])

    users = User.objects.select_related('department').all()

    for user in users:
        tasks = Task.objects.filter(executor=user, is_archived=False)

        sheet.append([
            user.username,
            user.email,
            user.department.name if user.department else 'Без отдела',
            tasks.count(),
            tasks.filter(status=Task.Status.TODO).count(),
            tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            tasks.filter(status=Task.Status.DONE).count(),
            tasks.filter(
                due_date__lt=now,
                status__in=[Task.Status.TODO, Task.Status.IN_PROGRESS],
            ).count(),
        ])

    for cell in sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(sheet)

    filename = f'executors_report_{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'
    return prepare_xlsx_response(workbook, filename)

@login_required
def projects_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    workbook = Workbook()

    # Лист 1: сводка
    summary_sheet = workbook.active
    summary_sheet.title = 'Проекты'

    summary_sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    summary_sheet.append([])

    summary_sheet.append([
        'Проект',
        'Владелец',
        'Всего задач',
        'К выполнению',
        'В работе',
        'Выполнено',
        'Просрочено',
    ])

    projects = Project.objects.select_related('owner').all()

    for project in projects:
        tasks = Task.objects.filter(project=project, is_archived=False)

        summary_sheet.append([
            project.name,
            project.owner.username,
            tasks.count(),
            tasks.filter(status=Task.Status.TODO).count(),
            tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            tasks.filter(status=Task.Status.DONE).count(),
            tasks.filter(
                due_date__lt=now,
                status__in=[Task.Status.TODO, Task.Status.IN_PROGRESS],
            ).count(),
        ])

    for cell in summary_sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(summary_sheet)

    # Лист 2: задачи по проектам
    tasks_sheet = workbook.create_sheet(title='Задачи проектов')

    tasks_sheet.append([
        'ID',
        'Название',
        'Ссылка',
        'Проект',
        'Статус',
        'Приоритет',
        'Постановщик',
        'Исполнитель',
        'Дедлайн',
        'Просрочена',
        'Создана',
        'Обновлена',
    ])

    tasks = Task.objects.select_related(
        'creator',
        'executor',
        'project',
    ).filter(
        is_archived=False,
        project__isnull=False,
    ).order_by('project__name', '-created_at')

    for task in tasks:
        link = get_task_url(request, task)

        tasks_sheet.append([
            task.pk,
            task.title,
            link,
            task.project.name if task.project else 'Без проекта',
            task.get_status_display(),
            task.get_priority_display(),
            task.creator.username,
            task.executor.username if task.executor else 'Не назначен',
            task.due_date.strftime('%d.%m.%Y %H:%M') if task.due_date else 'Не указан',
            'Да' if task.is_overdue else 'Нет',
            task.created_at.strftime('%d.%m.%Y %H:%M'),
            task.updated_at.strftime('%d.%m.%Y %H:%M'),
        ])

        row_num = tasks_sheet.max_row
        link_cell = tasks_sheet.cell(row=row_num, column=3)
        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'

    style_sheet(tasks_sheet)

    filename = f'projects_report_{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'
    return prepare_xlsx_response(workbook, filename)

@login_required
def departments_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    workbook = Workbook()

    # Лист 1: сводка
    summary_sheet = workbook.active
    summary_sheet.title = 'Отделы'

    summary_sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    summary_sheet.append([])

    summary_sheet.append([
        'Отдел',
        'Пользователей',
        'Всего задач',
        'К выполнению',
        'В работе',
        'Выполнено',
        'Просрочено',
    ])

    departments = Department.objects.prefetch_related('users').all()

    for department in departments:
        tasks = Task.objects.filter(
            executor__department=department,
            is_archived=False,
        )

        summary_sheet.append([
            department.name,
            department.users.count(),
            tasks.count(),
            tasks.filter(status=Task.Status.TODO).count(),
            tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            tasks.filter(status=Task.Status.DONE).count(),
            tasks.filter(
                due_date__lt=now,
                status__in=[Task.Status.TODO, Task.Status.IN_PROGRESS],
            ).count(),
        ])

    for cell in summary_sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(summary_sheet)

    # Лист 2: задачи по отделам
    tasks_sheet = workbook.create_sheet(title='Задачи отделов')

    tasks_sheet.append([
        'ID',
        'Название',
        'Ссылка',
        'Отдел',
        'Исполнитель',
        'Email исполнителя',
        'Проект',
        'Статус',
        'Приоритет',
        'Дедлайн',
        'Просрочена',
        'Создана',
        'Обновлена',
    ])

    tasks = Task.objects.select_related(
        'creator',
        'executor',
        'executor__department',
        'project',
    ).filter(
        is_archived=False,
        executor__department__isnull=False,
    ).order_by('executor__department__name', 'executor__username', '-created_at')

    for task in tasks:
        link = get_task_url(request, task)

        tasks_sheet.append([
            task.pk,
            task.title,
            link,
            task.executor.department.name if task.executor and task.executor.department else 'Без отдела',
            task.executor.username if task.executor else 'Не назначен',
            task.executor.email if task.executor else '',
            task.project.name if task.project else 'Без проекта',
            task.get_status_display(),
            task.get_priority_display(),
            task.due_date.strftime('%d.%m.%Y %H:%M') if task.due_date else 'Не указан',
            'Да' if task.is_overdue else 'Нет',
            task.created_at.strftime('%d.%m.%Y %H:%M'),
            task.updated_at.strftime('%d.%m.%Y %H:%M'),
        ])

        row_num = tasks_sheet.max_row
        link_cell = tasks_sheet.cell(row=row_num, column=3)
        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'

    style_sheet(tasks_sheet)

    filename = f'departments_report_{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'
    return prepare_xlsx_response(workbook, filename)

@login_required
def workload_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    workbook = Workbook()

    # ---------------------------
    # Лист 1: сводка по нагрузке
    # ---------------------------
    summary_sheet = workbook.active
    summary_sheet.title = 'Workload Report'

    summary_sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    summary_sheet.append([])

    summary_sheet.append([
        'Исполнитель',
        'Email',
        'Отдел',
        'Активных задач',
        'К выполнению',
        'В работе',
        'Высокий приоритет',
        'Просрочено',
        'Нагрузка',
    ])

    users = User.objects.select_related('department').all()

    for user in users:
        tasks = Task.objects.filter(
            executor=user,
            is_archived=False,
        )

        active_tasks = tasks.filter(
            status__in=[
                Task.Status.TODO,
                Task.Status.IN_PROGRESS,
            ]
        )

        active_count = active_tasks.count()

        if active_count >= 10:
            workload_level = 'Высокая'
        elif active_count >= 5:
            workload_level = 'Средняя'
        else:
            workload_level = 'Низкая'

        summary_sheet.append([
            user.username,
            user.email,
            user.department.name if user.department else 'Без отдела',

            active_count,

            active_tasks.filter(
                status=Task.Status.TODO
            ).count(),

            active_tasks.filter(
                status=Task.Status.IN_PROGRESS
            ).count(),

            active_tasks.filter(
                priority=Task.Priority.HIGH
            ).count(),

            active_tasks.filter(
                due_date__lt=now,
            ).count(),

            workload_level,
        ])

    for cell in summary_sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(summary_sheet)

    # ---------------------------
    # Лист 2: подробные задачи
    # ---------------------------
    tasks_sheet = workbook.create_sheet(
        title='Workload Tasks'
    )

    tasks_sheet.append([
        'ID',
        'Название',
        'Ссылка',
        'Исполнитель',
        'Email',
        'Отдел',
        'Статус',
        'Приоритет',
        'Проект',
        'Дедлайн',
        'Просрочена',
        'Создана',
        'Обновлена',
    ])

    tasks = Task.objects.select_related(
        'executor',
        'executor__department',
        'project',
    ).filter(
        is_archived=False,
        status__in=[
            Task.Status.TODO,
            Task.Status.IN_PROGRESS,
        ]
    ).order_by(
        'executor__username',
        '-priority',
        'due_date',
    )

    for task in tasks:

        link = request.build_absolute_uri(
            reverse(
                'tasks:detail',
                kwargs={'pk': task.pk}
            )
        )

        tasks_sheet.append([
            task.pk,

            task.title,

            link,

            task.executor.username
            if task.executor else 'Не назначен',

            task.executor.email
            if task.executor else '',

            task.executor.department.name
            if task.executor and task.executor.department
            else 'Без отдела',

            task.get_status_display(),

            task.get_priority_display(),

            task.project.name
            if task.project else 'Без проекта',

            task.due_date.strftime('%d.%m.%Y %H:%M')
            if task.due_date else 'Не указан',

            'Да' if task.is_overdue else 'Нет',

            task.created_at.strftime('%d.%m.%Y %H:%M'),

            task.updated_at.strftime('%d.%m.%Y %H:%M'),
        ])

        row_num = tasks_sheet.max_row

        link_cell = tasks_sheet.cell(
            row=row_num,
            column=3
        )

        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'

    for cell in tasks_sheet[1]:
        cell.font = Font(bold=True)

    style_sheet(tasks_sheet)

    # ---------------------------
    # Отдача файла
    # ---------------------------
    filename = (
        f'workload_report_'
        f'{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'
    )

    return prepare_xlsx_response(
        workbook,
        filename
    )

@login_required
def priority_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    workbook = Workbook()

    # ---------------------------
    # Лист 1: сводка по приоритетам
    # ---------------------------
    summary_sheet = workbook.active
    summary_sheet.title = 'Priority Report'

    summary_sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    summary_sheet.append([])

    summary_sheet.append([
        'Приоритет',
        'Всего задач',
        'Активных',
        'К выполнению',
        'В работе',
        'Выполнено',
        'Просрочено',
    ])

    priorities = [
        Task.Priority.HIGH,
        Task.Priority.MEDIUM,
        Task.Priority.LOW,
    ]

    for priority in priorities:
        tasks = Task.objects.filter(
            priority=priority,
            is_archived=False,
        )

        active_tasks = tasks.filter(
            status__in=[
                Task.Status.TODO,
                Task.Status.IN_PROGRESS,
            ]
        )

        summary_sheet.append([
            Task.Priority(priority).label,
            tasks.count(),
            active_tasks.count(),
            tasks.filter(status=Task.Status.TODO).count(),
            tasks.filter(status=Task.Status.IN_PROGRESS).count(),
            tasks.filter(status=Task.Status.DONE).count(),
            active_tasks.filter(due_date__lt=now).count(),
        ])

    for cell in summary_sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(summary_sheet)

    # ---------------------------
    # Лист 2: задачи подробно
    # ---------------------------
    tasks_sheet = workbook.create_sheet(title='Priority Tasks')

    tasks_sheet.append([
        'ID',
        'Название',
        'Ссылка',
        'Приоритет',
        'Статус',
        'Постановщик',
        'Исполнитель',
        'Проект',
        'Дедлайн',
        'Просрочена',
        'Создана',
        'Обновлена',
    ])

    tasks = Task.objects.select_related(
        'creator',
        'executor',
        'project',
    ).filter(
        is_archived=False,
    ).order_by(
        'priority',
        'due_date',
        '-created_at',
    )

    for task in tasks:
        link = request.build_absolute_uri(
            reverse(
                'tasks:detail',
                kwargs={'pk': task.pk}
            )
        )

        tasks_sheet.append([
            task.pk,
            task.title,
            link,
            task.get_priority_display(),
            task.get_status_display(),
            task.creator.username,
            task.executor.username if task.executor else 'Не назначен',
            task.project.name if task.project else 'Без проекта',
            task.due_date.strftime('%d.%m.%Y %H:%M') if task.due_date else 'Не указан',
            'Да' if task.is_overdue else 'Нет',
            task.created_at.strftime('%d.%m.%Y %H:%M'),
            task.updated_at.strftime('%d.%m.%Y %H:%M'),
        ])

        row_num = tasks_sheet.max_row
        link_cell = tasks_sheet.cell(row=row_num, column=3)
        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'

    for cell in tasks_sheet[1]:
        cell.font = Font(bold=True)

    style_sheet(tasks_sheet)

    filename = f'priority_report_{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'

    return prepare_xlsx_response(workbook, filename)

@login_required
def velocity_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    last_7_days = now - timedelta(days=7)
    last_30_days = now - timedelta(days=30)

    workbook = Workbook()

    # ---------------------------
    # Лист 1: Velocity по исполнителям
    # ---------------------------
    summary_sheet = workbook.active
    summary_sheet.title = 'Velocity Report'

    summary_sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    summary_sheet.append([])

    summary_sheet.append([
        'Исполнитель',
        'Email',
        'Отдел',
        'Закрыто за 7 дней',
        'Закрыто за 30 дней',
        'Всего закрыто',
        'Закрыто с просрочкой',
    ])

    users = User.objects.select_related('department').all()

    for user in users:
        completed_tasks = Task.objects.filter(
            executor=user,
            status=Task.Status.DONE,
            completed_at__isnull=False,
            is_archived=False,
        )

        overdue_completed = completed_tasks.filter(
            due_date__isnull=False,
            completed_at__gt=models.F('due_date'),
        ).count()

        summary_sheet.append([
            user.username,
            user.email,
            user.department.name if user.department else 'Без отдела',
            completed_tasks.filter(completed_at__gte=last_7_days).count(),
            completed_tasks.filter(completed_at__gte=last_30_days).count(),
            completed_tasks.count(),
            overdue_completed,
        ])

    for cell in summary_sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(summary_sheet)

    # ---------------------------
    # Лист 2: Velocity по дням
    # ---------------------------
    days_sheet = workbook.create_sheet(title='Velocity по дням')

    days_sheet.append([
        'Дата',
        'Закрыто задач',
    ])

    for i in range(29, -1, -1):
        day = now.date() - timedelta(days=i)

        completed_count = Task.objects.filter(
            status=Task.Status.DONE,
            completed_at__date=day,
            completed_at__isnull=False,
            is_archived=False,
        ).count()

        days_sheet.append([
            day.strftime('%d.%m.%Y'),
            completed_count,
        ])

    for cell in days_sheet[1]:
        cell.font = Font(bold=True)

    style_sheet(days_sheet)

    # ---------------------------
    # Лист 3: Закрытые задачи подробно
    # ---------------------------
    tasks_sheet = workbook.create_sheet(title='Closed Tasks')

    tasks_sheet.append([
        'ID',
        'Название',
        'Ссылка',
        'Исполнитель',
        'Email исполнителя',
        'Отдел',
        'Проект',
        'Приоритет',
        'Дедлайн',
        'Дата закрытия',
        'Закрыта с просрочкой',
        'Создана',
    ])

    tasks = Task.objects.select_related(
        'executor',
        'executor__department',
        'project',
    ).filter(
        status=Task.Status.DONE,
        completed_at__isnull=False,
        is_archived=False,
    ).order_by('-completed_at')

    for task in tasks:
        link = request.build_absolute_uri(
            reverse('tasks:detail', kwargs={'pk': task.pk})
        )

        is_overdue_completed = (
            task.due_date is not None
            and task.completed_at is not None
            and task.completed_at > task.due_date
        )

        tasks_sheet.append([
            task.pk,
            task.title,
            link,
            task.executor.username if task.executor else 'Не назначен',
            task.executor.email if task.executor else '',
            task.executor.department.name
            if task.executor and task.executor.department
            else 'Без отдела',
            task.project.name if task.project else 'Без проекта',
            task.get_priority_display(),
            task.due_date.strftime('%d.%m.%Y %H:%M') if task.due_date else 'Не указан',
            task.completed_at.strftime('%d.%m.%Y %H:%M'),
            'Да' if is_overdue_completed else 'Нет',
            task.created_at.strftime('%d.%m.%Y %H:%M'),
        ])

        row_num = tasks_sheet.max_row
        link_cell = tasks_sheet.cell(row=row_num, column=3)
        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'

    for cell in tasks_sheet[1]:
        cell.font = Font(bold=True)

    style_sheet(tasks_sheet)

    filename = f'velocity_report_{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'

    return prepare_xlsx_response(workbook, filename)

@login_required
def sla_report_xlsx(request):
    now = timezone.now()
    downloaded_at = now.strftime('%d.%m.%Y %H:%M')

    workbook = Workbook()

    # ---------------------------
    # Лист 1: SLA по исполнителям
    # ---------------------------
    summary_sheet = workbook.active
    summary_sheet.title = 'SLA Report'

    summary_sheet.append([
        'Дата скачивания',
        downloaded_at,
    ])

    summary_sheet.append([])

    summary_sheet.append([
        'Исполнитель',
        'Email',
        'Отдел',
        'Закрыто задач',
        'Вовремя',
        'С просрочкой',
        'SLA %',
        'Уровень SLA',
    ])

    users = User.objects.select_related('department').all()

    for user in users:
        completed_tasks = Task.objects.filter(
            executor=user,
            status=Task.Status.DONE,
            completed_at__isnull=False,
            due_date__isnull=False,
            is_archived=False,
        )

        total_completed = completed_tasks.count()

        completed_on_time = completed_tasks.filter(
            completed_at__lte=models.F('due_date')
        ).count()

        completed_overdue = completed_tasks.filter(
            completed_at__gt=models.F('due_date')
        ).count()

        if total_completed > 0:
            sla_percent = round((completed_on_time / total_completed) * 100, 1)
        else:
            sla_percent = 0

        if sla_percent >= 90:
            sla_level = 'Высокий'
        elif sla_percent >= 70:
            sla_level = 'Средний'
        else:
            sla_level = 'Низкий'

        summary_sheet.append([
            user.username,
            user.email,
            user.department.name if user.department else 'Без отдела',
            total_completed,
            completed_on_time,
            completed_overdue,
            sla_percent,
            sla_level,
        ])

    for cell in summary_sheet[3]:
        cell.font = Font(bold=True)

    style_sheet(summary_sheet)

    # ---------------------------
    # Лист 2: SLA сводка для pie chart
    # ---------------------------
    chart_sheet = workbook.create_sheet(title='SLA Pie Data')

    chart_sheet.append([
        'Категория',
        'Количество задач',
    ])

    total_on_time = Task.objects.filter(
        status=Task.Status.DONE,
        completed_at__isnull=False,
        due_date__isnull=False,
        completed_at__lte=models.F('due_date'),
        is_archived=False,
    ).count()

    total_overdue = Task.objects.filter(
        status=Task.Status.DONE,
        completed_at__isnull=False,
        due_date__isnull=False,
        completed_at__gt=models.F('due_date'),
        is_archived=False,
    ).count()

    chart_sheet.append([
        'Вовремя',
        total_on_time,
    ])

    chart_sheet.append([
        'С просрочкой',
        total_overdue,
    ])

    for cell in chart_sheet[1]:
        cell.font = Font(bold=True)

    style_sheet(chart_sheet)

    # ---------------------------
    # Лист 3: задачи подробно
    # ---------------------------
    tasks_sheet = workbook.create_sheet(title='SLA Tasks')

    tasks_sheet.append([
        'ID',
        'Название',
        'Ссылка',
        'Исполнитель',
        'Email исполнителя',
        'Отдел',
        'Проект',
        'Приоритет',
        'Дедлайн',
        'Дата закрытия',
        'SLA выполнен',
        'Создана',
        'Обновлена',
    ])

    tasks = Task.objects.select_related(
        'executor',
        'executor__department',
        'project',
    ).filter(
        status=Task.Status.DONE,
        completed_at__isnull=False,
        due_date__isnull=False,
        is_archived=False,
    ).order_by('-completed_at')

    for task in tasks:
        link = request.build_absolute_uri(
            reverse('tasks:detail', kwargs={'pk': task.pk})
        )

        sla_done = (
            task.completed_at is not None
            and task.due_date is not None
            and task.completed_at <= task.due_date
        )

        tasks_sheet.append([
            task.pk,
            task.title,
            link,
            task.executor.username if task.executor else 'Не назначен',
            task.executor.email if task.executor else '',
            task.executor.department.name
            if task.executor and task.executor.department
            else 'Без отдела',
            task.project.name if task.project else 'Без проекта',
            task.get_priority_display(),
            task.due_date.strftime('%d.%m.%Y %H:%M'),
            task.completed_at.strftime('%d.%m.%Y %H:%M'),
            'Да' if sla_done else 'Нет',
            task.created_at.strftime('%d.%m.%Y %H:%M'),
            task.updated_at.strftime('%d.%m.%Y %H:%M'),
        ])

        row_num = tasks_sheet.max_row
        link_cell = tasks_sheet.cell(row=row_num, column=3)
        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'

    for cell in tasks_sheet[1]:
        cell.font = Font(bold=True)

    style_sheet(tasks_sheet)

    filename = f'sla_report_{now.strftime("%Y_%m_%d_%H_%M")}.xlsx'

    return prepare_xlsx_response(workbook, filename)